#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador do PLANO DE ESTOQUE ESTRATÉGICO
=======================================

Lê o arquivo de origem (Estoque_Claude.xlsx) e produz um workbook com:

  1. Análise de demanda   -> vendas mensais/semestrais, ranking, demanda
                             DIRETA (vendas do próprio item), INDIRETA
                             (item como componente de estrutura/BOM) e TOTAL.
  2. Indicadores de estoque -> Safety Stock e Reorder Point pela formulação
                             estatística clássica:
                                 SS  = Z * sigma * sqrt(LT)
                                 ROP = (media * LT) + SS
                             com Lead Time vindo do Cadastro (coluna "Dias
                             Entrega") convertido para meses.

Este gerador substitui/aprimora o modelo de referência da aba "JHG", que
usava buckets trimestrais e uma fórmula de safety stock ad-hoc. Aqui a
mesma decomposição direta+indireta é mantida, porém sobre uma série
mensal completa e com a fórmula estatística pedida.

Uso:
    python3 gerar_plano_estoque.py [origem.xlsx] [saida.xlsx]
"""

import sys
import math
import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# --------------------------------------------------------------------------
# Parâmetros de negócio (editáveis também na aba "Parâmetros" do resultado)
# --------------------------------------------------------------------------
NIVEL_SERVICO = 0.95        # nível de serviço alvo
Z_SCORE = 1.645             # z correspondente a 95% (one-tailed)
DIAS_POR_MES = 30.0         # conversão de lead time (dias -> meses)
TOP_RANKING = 50            # tamanho do ranking destacado

# Índices de coluna (0-based) nas abas de origem
V_PRODUTO, V_DESC, V_QTD, V_DTEMISSAO = 5, 6, 7, 13   # aba Vendas
C_COD, C_DESC, C_TIPO, C_UN, C_LT = 0, 1, 2, 3, 4      # aba Cadastro
E_PARENT, E_COMP, E_DESC, E_SEQ, E_QTD = 0, 1, 2, 3, 4  # aba Estruturas

SHEET_CAD = "1-Cadastro de Produtos"
SHEET_EST = "Estruturas"
SHEET_VEN = "Vendas"


# --------------------------------------------------------------------------
# Leitura da base
# --------------------------------------------------------------------------
def ler_base(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

    # ---- Cadastro: código -> {desc, tipo, un, lead_time_dias}
    cadastro = {}
    ws = wb[SHEET_CAD]
    for r in ws.iter_rows(min_row=3, values_only=True):
        cod = r[C_COD]
        if cod is None:
            continue
        try:
            lt = float(r[C_LT]) if r[C_LT] is not None else None
        except (TypeError, ValueError):
            lt = None
        cadastro[cod] = {
            "desc": r[C_DESC] or "",
            "tipo": r[C_TIPO] or "",
            "un": r[C_UN] or "",
            "lt_dias": lt,
        }

    # ---- Estruturas (BOM): lista de (parent, componente, qtd)
    bom = []
    ws = wb[SHEET_EST]
    for r in ws.iter_rows(min_row=3, values_only=True):
        par, comp, qtd = r[E_PARENT], r[E_COMP], r[E_QTD]
        if par is None or comp is None:
            continue
        try:
            q = float(qtd) if qtd is not None else 1.0
        except (TypeError, ValueError):
            q = 1.0
        bom.append((par, comp, q, r[E_DESC] or ""))

    # ---- Vendas: agrega quantidade por (produto, mês)
    vendas_mensal = defaultdict(lambda: defaultdict(float))  # produto -> mes(str) -> qtd
    desc_venda = {}
    datas = []
    ws = wb[SHEET_VEN]
    for r in ws.iter_rows(min_row=2, values_only=True):
        prod = r[V_PRODUTO]
        qtd = r[V_QTD]
        dt = r[V_DTEMISSAO]
        if prod is None or not isinstance(dt, datetime.datetime):
            continue
        try:
            q = float(qtd)
        except (TypeError, ValueError):
            continue
        mes = f"{dt.year:04d}-{dt.month:02d}"
        vendas_mensal[prod][mes] += q
        datas.append(dt)
        if prod not in desc_venda and r[V_DESC]:
            desc_venda[prod] = r[V_DESC]

    wb.close()
    return cadastro, bom, vendas_mensal, desc_venda, datas


def lista_meses(datas):
    """Todos os meses (YYYY-MM) da janela contínua min..max das vendas."""
    dmin, dmax = min(datas), max(datas)
    meses = []
    y, m = dmin.year, dmin.month
    while (y, m) <= (dmax.year, dmax.month):
        meses.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return meses


def semestre_de(mes):
    y, m = mes.split("-")
    return f"{y}-S{1 if int(m) <= 6 else 2}"


# --------------------------------------------------------------------------
# Cálculo de demanda direta / indireta / total
# --------------------------------------------------------------------------
def calcular(cadastro, bom, vendas_mensal, desc_venda, meses):
    meses_set = set(meses)

    # Demanda DIRETA por (produto, mês) — somente meses da janela
    direta = defaultdict(lambda: defaultdict(float))
    for prod, mm in vendas_mensal.items():
        for mes, q in mm.items():
            if mes in meses_set:
                direta[prod][mes] += q

    # Demanda INDIRETA: para cada linha do BOM (pai -> componente x qtd),
    # o consumo do componente segue as vendas diretas do produto-pai.
    indireta = defaultdict(lambda: defaultdict(float))
    for par, comp, q, _desc in bom:
        vendas_pai = direta.get(par)
        if not vendas_pai:
            continue
        for mes, vq in vendas_pai.items():
            indireta[comp][mes] += vq * q

    # Universo de itens: vendidos + componentes de estrutura
    itens = set(direta.keys()) | set(indireta.keys())

    resultados = {}
    n_meses = len(meses)
    for it in itens:
        serie_dir = [direta[it].get(m, 0.0) for m in meses]
        serie_ind = [indireta[it].get(m, 0.0) for m in meses]
        serie_tot = [d + i for d, i in zip(serie_dir, serie_ind)]

        tot_dir = sum(serie_dir)
        tot_ind = sum(serie_ind)
        tot_tot = sum(serie_tot)

        media = tot_tot / n_meses if n_meses else 0.0
        # desvio padrão amostral (ddof=1) da demanda mensal total
        if n_meses > 1:
            var = sum((x - media) ** 2 for x in serie_tot) / (n_meses - 1)
            desvio = math.sqrt(var)
        else:
            desvio = 0.0

        info = cadastro.get(it, {})
        lt_dias = info.get("lt_dias")
        desc = info.get("desc") or desc_venda.get(it, "")

        resultados[it] = {
            "desc": desc,
            "tipo": info.get("tipo", ""),
            "un": info.get("un", ""),
            "lt_dias": lt_dias,
            "serie_dir": serie_dir,
            "serie_ind": serie_ind,
            "serie_tot": serie_tot,
            "tot_dir": tot_dir,
            "tot_ind": tot_ind,
            "tot_tot": tot_tot,
            "media_mensal": media,
            "desvio_mensal": desvio,
        }

    return resultados, direta, indireta


# --------------------------------------------------------------------------
# Escrita do workbook de saída
# --------------------------------------------------------------------------
# paleta
AZUL = "1F3864"
AZUL_MED = "2E5496"
AZUL_CLARO = "D6E0F0"
CINZA = "F2F2F2"
VERDE = "375623"
VERDE_CLARO = "E2EFDA"
LARANJA = "C55A11"
BRANCO = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDA = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(ws, row, headers, fill=AZUL, start_col=1):
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + j, value=h)
        c.font = Font(bold=True, color=BRANCO, size=10)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA
    ws.row_dimensions[row].height = 30


def titulo(ws, texto, sub=None):
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = texto
    c.font = Font(bold=True, size=16, color=AZUL)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    if sub:
        ws.merge_cells("A2:H2")
        c2 = ws["A2"]
        c2.value = sub
        c2.font = Font(italic=True, size=10, color="595959")


def escrever(caminho_saida, origem, cadastro, bom, vendas_mensal,
             desc_venda, meses, resultados, direta, indireta):
    wb = openpyxl.Workbook()
    # força qualquer visualizador a recalcular as fórmulas ao abrir
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    semestres = []
    for m in meses:
        s = semestre_de(m)
        if s not in semestres:
            semestres.append(s)

    itens_ordenados = sorted(resultados.keys(),
                             key=lambda k: resultados[k]["tot_tot"], reverse=True)

    # ================= 1. LEIA-ME =================
    ws = wb.active
    ws.title = "Leia-me"
    titulo(ws, "PLANO DE ESTOQUE ESTRATÉGICO",
           "Análise de demanda, Safety Stock e Ponto de Reposição")
    linhas = [
        ("", ""),
        ("COMO ESTE MODELO FUNCIONA", ""),
        ("Fonte", f"Gerado a partir de '{origem}'."),
        ("Base de data", "A demanda é datada pela 'DT Emissao' do pedido (sinal "
                         "de demanda do cliente). Para usar a data de faturamento/"
                         "entrega, altere V_DTEMISSAO no gerador."),
        ("Janela de análise", f"{meses[0]} a {meses[-1]}  ({len(meses)} meses)"),
        ("Itens analisados", f"{len(resultados)} (vendidos + componentes de estrutura)"),
        ("", ""),
        ("DEMANDA", ""),
        ("Demanda DIRETA", "Vendas do próprio item (aba Vendas, por mês)."),
        ("Demanda INDIRETA", "Consumo do item quando ele é componente de outro "
                             "produto estruturado (BOM). Para cada estrutura "
                             "pai->componente, o consumo segue as vendas do pai "
                             "multiplicadas pela quantidade da estrutura."),
        ("Demanda TOTAL", "Direta + Indireta, mês a mês."),
        ("", ""),
        ("INDICADORES DE ESTOQUE", ""),
        ("Demanda média (μ)", "Média da demanda mensal total na janela."),
        ("Desvio padrão (σ)", "Desvio padrão amostral da demanda mensal total."),
        ("Lead Time (LT)", "Aba Cadastro, coluna 'Dias Entrega'. Convertido para "
                           f"meses dividindo por {DIAS_POR_MES:.0f} dias/mês."),
        ("Z-Score", f"{Z_SCORE} (nível de serviço {NIVEL_SERVICO:.0%})."),
        ("", ""),
        ("FÓRMULAS", ""),
        ("Safety Stock", "SS = Z × σ × √(Lead Time em meses)"),
        ("Ponto de Reposição", "ROP = (μ × Lead Time em meses) + Safety Stock"),
        ("Demanda no Lead Time", "μ × Lead Time (em meses)"),
        ("", ""),
        ("PARÂMETROS AJUSTÁVEIS", "Veja a aba 'Parâmetros'. Ao alterar o Z-Score, "
                                  "as colunas de Safety Stock e Ponto de Reposição "
                                  "recalculam automaticamente."),
        ("", ""),
        ("ABAS", ""),
        ("Parâmetros", "Nível de serviço / Z-Score / dias por mês."),
        ("Análise de Demanda", "Ranking, demanda direta/indireta/total, médias."),
        ("Indicadores de Estoque", "Safety Stock e Ponto de Reposição por item."),
        ("Demanda Mensal (D+I)", "Matriz item × mês da demanda total."),
        ("Demanda Semestral", "Demanda total consolidada por semestre."),
        (f"Ranking Top {TOP_RANKING}", "Itens mais demandados (com gráfico)."),
        ("Uso em Estruturas (BOM)", "Estruturas em que cada item é componente."),
        ("Base Vendas / Cadastro / BOM", "Dados de origem para auditoria."),
    ]
    r = 3
    for a, b in linhas:
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        if a and not b:
            ca.font = Font(bold=True, color=AZUL_MED, size=11)
        else:
            ca.font = Font(bold=True, size=10)
        cb.font = Font(size=10)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90
    ws.sheet_view.showGridLines = False

    # ================= 2. PARÂMETROS =================
    wp = wb.create_sheet("Parâmetros")
    titulo(wp, "PARÂMETROS", "Edite as células amarelas — o modelo recalcula.")
    param_rows = [
        ("Nível de serviço", NIVEL_SERVICO, "0.0%"),
        ("Z-Score", Z_SCORE, "0.000"),
        ("Dias por mês", DIAS_POR_MES, "0"),
    ]
    hdr(wp, 4, ["Parâmetro", "Valor", ""], fill=AZUL_MED)
    amarelo = PatternFill("solid", fgColor="FFF2CC")
    r = 5
    param_cell = {}
    for nome, val, fmt in param_rows:
        wp.cell(row=r, column=1, value=nome).font = Font(bold=True, size=10)
        c = wp.cell(row=r, column=2, value=val)
        c.fill = amarelo
        c.number_format = fmt
        c.border = BORDA
        c.alignment = Alignment(horizontal="center")
        param_cell[nome] = f"Parâmetros!$B${r}"
        r += 1
    wp.cell(row=r + 1, column=1,
            value="Dica: nível de serviço 90%→Z=1.282, 95%→1.645, "
                  "97.5%→1.960, 99%→2.326.").font = Font(italic=True, size=9, color="808080")
    wp.column_dimensions["A"].width = 22
    wp.column_dimensions["B"].width = 14
    wp.sheet_view.showGridLines = False
    Z_REF = param_cell["Z-Score"]
    DPM_REF = param_cell["Dias por mês"]

    # ================= 3. ANÁLISE DE DEMANDA =================
    wa = wb.create_sheet("Análise de Demanda")
    titulo(wa, "ANÁLISE DE DEMANDA",
           f"Demanda direta, indireta e total por item — janela {meses[0]} a {meses[-1]}")
    cols = ["Rank", "Código", "Descrição", "Tipo", "Un", "Lead Time (dias)",
            "Demanda Direta (un)", "Demanda Indireta (un)", "Demanda Total (un)",
            "% do Total", "Média Mensal", "Média Semestral", "Desvio Padrão Mensal"]
    hdr(wa, 4, cols)
    total_geral = sum(resultados[i]["tot_tot"] for i in itens_ordenados) or 1.0
    r = 5
    for rank, it in enumerate(itens_ordenados, 1):
        d = resultados[it]
        vals = [
            rank, it, d["desc"], d["tipo"], d["un"], d["lt_dias"],
            round(d["tot_dir"], 2), round(d["tot_ind"], 2), round(d["tot_tot"], 2),
            d["tot_tot"] / total_geral,
            round(d["media_mensal"], 3),
            round(d["tot_tot"] / len(semestres), 3),
            round(d["desvio_mensal"], 3),
        ]
        for j, v in enumerate(vals, 1):
            c = wa.cell(row=r, column=j, value=v)
            c.border = BORDA
            c.font = Font(size=9)
            if j == 10:
                c.number_format = "0.0%"
            elif j in (7, 8, 9, 11, 12, 13):
                c.number_format = "#,##0.00"
        if rank % 2 == 0:
            for j in range(1, len(cols) + 1):
                wa.cell(row=r, column=j).fill = PatternFill("solid", fgColor=CINZA)
        r += 1
    widths = [6, 14, 42, 6, 5, 12, 15, 16, 15, 10, 12, 13, 16]
    for j, w in enumerate(widths, 1):
        wa.column_dimensions[get_column_letter(j)].width = w
    wa.freeze_panes = "A5"
    wa.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{r-1}"
    wa.sheet_view.showGridLines = False

    # ================= 4. INDICADORES DE ESTOQUE =================
    wi = wb.create_sheet("Indicadores de Estoque")
    titulo(wi, "INDICADORES DE ESTOQUE",
           "Safety Stock = Z×σ×√LT   |   Ponto de Reposição = (μ×LT) + SS")
    cols = ["Código", "Descrição", "Tipo", "Lead Time (dias)", "Lead Time (meses)",
            "Demanda Média Mensal (μ)", "Desvio Padrão Mensal (σ)", "CV (σ/μ)",
            "Z-Score", "Demanda no Lead Time", "Safety Stock (un)",
            "Ponto de Reposição (un)"]
    hdr(wi, 4, cols, fill=VERDE)
    r = 5
    for it in itens_ordenados:
        d = resultados[it]
        lt = d["lt_dias"]
        mu = round(d["media_mensal"], 6)
        sg = round(d["desvio_mensal"], 6)
        # colunas: A cod B desc C tipo D lt_dias E lt_meses F mu G sigma
        #          H cv I z J dem_lt K ss L rop
        cA = wi.cell(row=r, column=1, value=it)
        cB = wi.cell(row=r, column=2, value=d["desc"])
        cC = wi.cell(row=r, column=3, value=d["tipo"])
        cD = wi.cell(row=r, column=4, value=lt)
        cE = wi.cell(row=r, column=5,
                     value=(f"=IF(D{r}=\"\",\"\",D{r}/{DPM_REF})"))
        cF = wi.cell(row=r, column=6, value=mu)
        cG = wi.cell(row=r, column=7, value=sg)
        cH = wi.cell(row=r, column=8, value=f"=IF(F{r}=0,\"\",G{r}/F{r})")
        cI = wi.cell(row=r, column=9, value=f"={Z_REF}")
        cJ = wi.cell(row=r, column=10, value=f"=IF(E{r}=\"\",\"\",F{r}*E{r})")
        cK = wi.cell(row=r, column=11,
                     value=f"=IF(E{r}=\"\",\"\",I{r}*G{r}*SQRT(E{r}))")
        cL = wi.cell(row=r, column=12,
                     value=f"=IF(E{r}=\"\",\"\",J{r}+K{r})")
        for c in (cA, cB, cC, cD, cE, cF, cG, cH, cI, cJ, cK, cL):
            c.border = BORDA
            c.font = Font(size=9)
        for col in (5, 6, 7, 10):
            wi.cell(row=r, column=col).number_format = "#,##0.000"
        wi.cell(row=r, column=8).number_format = "0.00"
        wi.cell(row=r, column=9).number_format = "0.000"
        for col in (11, 12):
            cc = wi.cell(row=r, column=col)
            cc.number_format = "#,##0.00"
            cc.font = Font(size=9, bold=True, color=VERDE if col == 12 else LARANJA)
        wi.cell(row=r, column=11).fill = PatternFill("solid", fgColor="FCE4D6")
        wi.cell(row=r, column=12).fill = PatternFill("solid", fgColor=VERDE_CLARO)
        r += 1
    widths = [14, 40, 6, 13, 15, 20, 20, 9, 9, 16, 14, 17]
    for j, w in enumerate(widths, 1):
        wi.column_dimensions[get_column_letter(j)].width = w
    wi.freeze_panes = "A5"
    wi.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{r-1}"
    wi.sheet_view.showGridLines = False

    # ================= 5. DEMANDA MENSAL (D+I) =================
    wm = wb.create_sheet("Demanda Mensal (D+I)")
    titulo(wm, "DEMANDA MENSAL (DIRETA + INDIRETA)",
           "Matriz item × mês da demanda total")
    head = ["Código", "Descrição"] + meses + ["Total"]
    hdr(wm, 4, head, fill=AZUL_MED)
    r = 5
    for it in itens_ordenados:
        d = resultados[it]
        wm.cell(row=r, column=1, value=it).font = Font(size=9)
        wm.cell(row=r, column=2, value=d["desc"]).font = Font(size=9)
        for j, v in enumerate(d["serie_tot"]):
            if v:
                c = wm.cell(row=r, column=3 + j, value=round(v, 2))
                c.number_format = "#,##0.##"
                c.font = Font(size=9)
        ct = wm.cell(row=r, column=3 + len(meses), value=round(d["tot_tot"], 2))
        ct.font = Font(size=9, bold=True)
        ct.number_format = "#,##0.##"
        for j in range(1, 3 + len(meses) + 1):
            wm.cell(row=r, column=j).border = BORDA
        r += 1
    wm.column_dimensions["A"].width = 14
    wm.column_dimensions["B"].width = 38
    for j in range(len(meses)):
        wm.column_dimensions[get_column_letter(3 + j)].width = 8
    wm.column_dimensions[get_column_letter(3 + len(meses))].width = 10
    wm.freeze_panes = "C5"
    wm.sheet_view.showGridLines = False

    # ================= 6. DEMANDA SEMESTRAL =================
    wsem = wb.create_sheet("Demanda Semestral")
    titulo(wsem, "DEMANDA SEMESTRAL (DIRETA + INDIRETA)",
           "Demanda total consolidada por semestre")
    head = ["Código", "Descrição"] + semestres + ["Total"]
    hdr(wsem, 4, head, fill=AZUL_MED)
    # mapa mês->indice semestre
    idx_sem = {s: k for k, s in enumerate(semestres)}
    r = 5
    for it in itens_ordenados:
        d = resultados[it]
        buckets = [0.0] * len(semestres)
        for m, v in zip(meses, d["serie_tot"]):
            buckets[idx_sem[semestre_de(m)]] += v
        wsem.cell(row=r, column=1, value=it).font = Font(size=9)
        wsem.cell(row=r, column=2, value=d["desc"]).font = Font(size=9)
        for j, v in enumerate(buckets):
            c = wsem.cell(row=r, column=3 + j, value=round(v, 2) if v else None)
            c.number_format = "#,##0.##"
            c.font = Font(size=9)
        ct = wsem.cell(row=r, column=3 + len(semestres), value=round(d["tot_tot"], 2))
        ct.font = Font(size=9, bold=True)
        ct.number_format = "#,##0.##"
        for j in range(1, 3 + len(semestres) + 1):
            wsem.cell(row=r, column=j).border = BORDA
        r += 1
    wsem.column_dimensions["A"].width = 14
    wsem.column_dimensions["B"].width = 40
    for j in range(len(semestres)):
        wsem.column_dimensions[get_column_letter(3 + j)].width = 11
    wsem.column_dimensions[get_column_letter(3 + len(semestres))].width = 11
    wsem.freeze_panes = "C5"
    wsem.sheet_view.showGridLines = False

    # ================= 7. RANKING TOP N =================
    wr = wb.create_sheet(f"Ranking Top {TOP_RANKING}")
    titulo(wr, f"RANKING — TOP {TOP_RANKING} ITENS POR DEMANDA TOTAL")
    hdr(wr, 4, ["Rank", "Código", "Descrição", "Demanda Total (un)",
                "Demanda Direta", "Demanda Indireta"], fill=LARANJA)
    r = 5
    for rank, it in enumerate(itens_ordenados[:TOP_RANKING], 1):
        d = resultados[it]
        row_vals = [rank, it, d["desc"], round(d["tot_tot"], 2),
                    round(d["tot_dir"], 2), round(d["tot_ind"], 2)]
        for j, v in enumerate(row_vals, 1):
            c = wr.cell(row=r, column=j, value=v)
            c.border = BORDA
            c.font = Font(size=9)
            if j >= 4:
                c.number_format = "#,##0.##"
        r += 1
    wr.column_dimensions["A"].width = 6
    wr.column_dimensions["B"].width = 14
    wr.column_dimensions["C"].width = 46
    for col in ("D", "E", "F"):
        wr.column_dimensions[col].width = 16
    wr.sheet_view.showGridLines = False
    # gráfico de barras top 15
    n_graf = min(15, TOP_RANKING)
    chart = BarChart()
    chart.type = "bar"
    chart.title = f"Top {n_graf} por Demanda Total"
    chart.height = 10
    chart.width = 22
    data = Reference(wr, min_col=4, min_row=4, max_row=4 + n_graf)
    cats = Reference(wr, min_col=2, min_row=5, max_row=4 + n_graf)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    wr.add_chart(chart, "H4")

    # ================= 8. USO EM ESTRUTURAS (BOM) =================
    wb_ = wb.create_sheet("Uso em Estruturas (BOM)")
    titulo(wb_, "USO EM ESTRUTURAS (BILL OF MATERIALS)",
           "Estruturas em que o item aparece como componente e a demanda "
           "indireta gerada")
    hdr(wb_, 4, ["Componente", "Descrição Componente", "Estrutura (Pai)",
                 "Qtd por Unidade", "Vendas do Pai (un, janela)",
                 "Demanda Indireta Gerada (un)"], fill=AZUL_MED)
    r = 5
    for par, comp, q, desc in sorted(bom, key=lambda x: (x[1], x[0])):
        vendas_pai = sum(direta.get(par, {}).values())
        wb_.cell(row=r, column=1, value=comp).font = Font(size=9)
        wb_.cell(row=r, column=2, value=desc).font = Font(size=9)
        wb_.cell(row=r, column=3, value=par).font = Font(size=9)
        wb_.cell(row=r, column=4, value=q).font = Font(size=9)
        wb_.cell(row=r, column=5, value=round(vendas_pai, 2)).font = Font(size=9)
        wb_.cell(row=r, column=6, value=round(vendas_pai * q, 2)).font = Font(size=9)
        for j in (4, 5, 6):
            wb_.cell(row=r, column=j).number_format = "#,##0.##"
        for j in range(1, 7):
            wb_.cell(row=r, column=j).border = BORDA
        r += 1
    for col, w in zip("ABCDEF", [16, 34, 18, 14, 20, 22]):
        wb_.column_dimensions[col].width = w
    wb_.freeze_panes = "A5"
    wb_.sheet_view.showGridLines = False

    # ================= 9. BASES DE ORIGEM (auditoria) =================
    # Base Vendas (colunas-chave)
    bv = wb.create_sheet("Base Vendas")
    hdr(bv, 1, ["Produto", "Descrição", "Mês", "Quantidade"], fill="808080")
    r = 2
    for prod, mm in sorted(vendas_mensal.items()):
        for mes, q in sorted(mm.items()):
            bv.cell(row=r, column=1, value=prod)
            bv.cell(row=r, column=2, value=desc_venda.get(prod, ""))
            bv.cell(row=r, column=3, value=mes)
            bv.cell(row=r, column=4, value=round(q, 2))
            r += 1
    for col, w in zip("ABCD", [14, 42, 10, 12]):
        bv.column_dimensions[col].width = w
    bv.freeze_panes = "A2"

    # Base Cadastro
    bc = wb.create_sheet("Base Cadastro")
    hdr(bc, 1, ["Código", "Descrição", "Tipo", "Un", "Lead Time (dias)"], fill="808080")
    r = 2
    for cod, info in sorted(cadastro.items()):
        bc.cell(row=r, column=1, value=cod)
        bc.cell(row=r, column=2, value=info["desc"])
        bc.cell(row=r, column=3, value=info["tipo"])
        bc.cell(row=r, column=4, value=info["un"])
        bc.cell(row=r, column=5, value=info["lt_dias"])
        r += 1
    for col, w in zip("ABCDE", [14, 44, 6, 5, 14]):
        bc.column_dimensions[col].width = w
    bc.freeze_panes = "A2"

    # Base BOM
    bb = wb.create_sheet("Base BOM")
    hdr(bb, 1, ["Estrutura (Pai)", "Componente", "Descrição", "Qtd"], fill="808080")
    r = 2
    for par, comp, q, desc in bom:
        bb.cell(row=r, column=1, value=par)
        bb.cell(row=r, column=2, value=comp)
        bb.cell(row=r, column=3, value=desc)
        bb.cell(row=r, column=4, value=q)
        r += 1
    for col, w in zip("ABCD", [16, 16, 40, 8]):
        bb.column_dimensions[col].width = w
    bb.freeze_panes = "A2"

    wb.save(caminho_saida)
    return len(resultados)


# --------------------------------------------------------------------------
def main():
    origem = sys.argv[1] if len(sys.argv) > 1 else "Estoque_Claude.xlsx"
    saida = sys.argv[2] if len(sys.argv) > 2 else "Plano_Estoque_Estrategico.xlsx"

    print(f"Lendo base: {origem}")
    cadastro, bom, vendas_mensal, desc_venda, datas = ler_base(origem)
    meses = lista_meses(datas)
    print(f"  Cadastro: {len(cadastro)} produtos | BOM: {len(bom)} linhas | "
          f"Vendas: {len(vendas_mensal)} produtos | Meses: {len(meses)}")

    resultados, direta, indireta = calcular(
        cadastro, bom, vendas_mensal, desc_venda, meses)

    n = escrever(saida, origem.split("/")[-1], cadastro, bom, vendas_mensal,
                 desc_venda, meses, resultados, direta, indireta)
    print(f"OK -> {saida}  ({n} itens)")


if __name__ == "__main__":
    main()
